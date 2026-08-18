"""
BOM(Excel) 与 PDF 原理图 逐器件核对工具
========================================
功能：

author : Jean 金成 QQ：173722238 
  1. 读取excel文件的任意列
  2. 精确解码 PDF 原理图文字（原理图的字体使用"字形编号+29=ASCII码"的自定义编码，
     需从内容流中还原，OCR 无法保证完整识别）
  3. 二者一一核对：一致(绿)/不一致(红) 全部列出
  4. 生成 Excel 报告(.xlsx) 与文本报告(.txt)

适用：PAT - 本项目仅针对该系列 PDF(SEYOND/SIPR 原理图导出格式)。
"""

import os  
import re   #正则表达式，用于字符串匹配，查找，替换，文本提取
import sys  # python 解释器，命令行参数，脚本控制，修改搜索路径
import time 
import threading # 多线程，开启多个字现成并发执行IO，
import traceback   # 大于异常堆栈和模块
from collections import defaultdict #把`collections`里面的`defaultdict`拿过来直接用

# ---------- 依赖检查 ----------
REQUIRED = []
try:
    import pikepdf
except ImportError:
    REQUIRED.append("pikepdf")
try:
    import pdfplumber
except ImportError:
    REQUIRED.append("pdfplumber")
try:
    import openpyxl
except ImportError:
    REQUIRED.append("openpyxl")

if REQUIRED or True:
    import pikepdf  # 再次导入确保错误清晰
    import pdfplumber
    import openpyxl

# ---------- PDF 解码 -------------------------------
DESIGNATOR_RE = re.compile(r"^[A-Za-z]{1,4}\d{1,4}(?:[A-Za-z]\d{1,2})?$")
PREFIXES_PARTS = set("CRLUDQJRTFXSWYKT")


def _build_code2glyph(fonts):
    """把 Type3 字体的 Differences 数组解析成 {字符码: 字形名}。"""
    code2glyph = {} # 定义字典，
    for fn in fonts.keys(): # fn 为字体字符串，fonts[fn] 为字体对象  
        f = fonts[fn]   # 赋值
        enc = f.get("/Encoding") # 获取字体编码对象
        diffs = None #数组为空
        if enc is not None and hasattr(enc, "get"):  #
            diffs = enc.get("/Differences", None)
        if diffs is None:
            continue
        try:  # 将diffs中的每个元素转换为字符串，并存储在arr列表中
            arr = [str(x) for x in diffs]  # 数组转换
        except Exception: 
            continue
        gmap = {} # 定义字典 gmap
        start = None
        names = [] # 列表 names 名称

        def flush():  # 定义函数
            nonlocal gmap, start, names #循环，将start和names中的元素添加到gmap字典中
            if start is None:
                return
            cur = start
            for g in names:
                gmap[cur] = g
                cur += 1
            start = None  # 处理完成，进行重置
            names = []

        for x in arr: 
            if re.fullmatch(r"-?\d+", x): # 正则表达式，判断元素是否整数（正负数字字符串）
                flush() #调用函数
                start = int(x) 
            else:
                names.append(x)
        flush() #调用，避免数组
        code2glyph[str(fn)] = gmap
    return code2glyph


def decode_page(pdf, pageno): 
    """完整解码一页：返回[(字符, x, y, 字号)]，自原点为页面左上，y 越上越小。"""
    from pikepdf import parse_content_stream #库pikepdf中的函数parse_content_stream，解析PDF内容流

    page = pdf.pages[pageno]  # 获取指定页码的页面对象
    c2g = _build_code2glyph(page["/Resources"]["/Font"])  #`c2g` = code‑to‑glyph：**字体名称 → {字节编码：字形名}**
    ops = list(parse_content_stream(page.Contents))
    tm = [1, 0, 0, 1, 0, 0]  #状态变量
    tl = 0.0 #原点00
    cur_font = None #当前字体，默认不使用
    cur_size = 1.0  #缩放系数默认1.0
    out = [] #输出列表，存放解析之后提出的文本内容
    for op in ops:
        name = str(op.operator) 
        ops_ = list(op.operands)
        if name == "BT":
            tm = [1, 0, 0, 1, 0, 0]
        elif name == "Tf":
            cur_font = str(ops_[0])
            try:
                cur_size = float(ops_[1])
            except Exception:
                pass
        elif name == "Tm":
            tm = [float(x) for x in ops_[:6]]
        elif name == "Td":
            tx, ty = float(ops_[0]), float(ops_[1])
            a, b, c, d, e, f = tm
            tm = [a, b, c, d, e + tx * a + ty * c, f + tx * b + ty * d]
        elif name == "T*":
            a, b, c, d, e, f = tm
            tm = [a, b, c, d, e, f + tl * d]
        elif name == "TL":
            tl = float(ops_[0])
        elif name in ("Tj", "'"):
            b = bytes(ops_[0]) if hasattr(ops_[0], "__bytes__") else b""
            gmap = c2g.get(cur_font, {})
            for byte in b:
                _append_char(out, gmap, byte, tm, cur_size)
        elif name == "TJ":
            for item in ops_[0]:
                if hasattr(item, "__bytes__"):
                    b = bytes(item)
                    gmap = c2g.get(cur_font, {})
                    for byte in b:
                        _append_char(out, gmap, byte, tm, cur_size)
    return out


def _append_char(out, gmap, byte, tm, cur_size): 
    glyph = gmap.get(byte)
    m = re.match(r"/g(\d+)$", glyph or "")
    num = int(m.group(1)) if m else None
    ch = chr(num + 29) if num is not None and 0 < num + 29 <= 0x10FFFF else "?"
    out.append((ch, tm[4], tm[5], cur_size))

# 根据传入的字节值（byte）在字形映射表（gmap）中查找对应的字形名称，解析出字形编号，
# 并将其转换为实际的字符，最后将该字符连同坐标与字号信息一起存入输出列表中。

def decode_pdf_blocks(pdf_path):
    """解码 PDF 全部页面，返回页面文字行(列表) 与 全部字符。
    页序 1..N。每行保留位置信息用于后续定位。
    """
    pdf = pikepdf.open(pdf_path)
    pages_words = {}     # page -> {y: [(word, x0, x1)]}  （已按行分词组）
    pages_text = {}      # page -> [行文字]
    page_count = len(pdf.pages)

    for pno in range(page_count):
        chars = decode_page(pdf, pno)
        # 与 pdfplumber 的精确 bbox 对齐（保障字符宽度/间距正确）
        boxes = _get_char_boxes(pdf_path, pno)
        if boxes and len(boxes) == len(chars):
            merged = []
            for (ch, x, y, sz), (x0, top, x1, bottom) in zip(chars, boxes):
                merged.append((ch, y, sz, x0, x1, top))
        else:
            merged = [(ch, y, sz, x, x, y) for (ch, x, y, sz) in chars]

        # 以 pdfplumber 的 top(同行近似)为行键
        lines = defaultdict(list)
        for item in merged: # 遍历合并后的字符信息
            ch, y, sz, x0, x1, top = item # 解包字符信息
            key = round(top, 7) # 将 top 坐标四舍五入到小数点后 7 位，作为行的键
            lines[key].append((ch, x0, x1, sz, top)) # 将字符信息添加到对应行的列表

        page_words = []  # 每页的行列表，每行包含字符及其位置信息
        page_text = []  # 每页的行文字列表，每行仅包含文本内容
        for yk in sorted(lines.keys()): # 按行键排序，确保行的顺序与 PDF 中的顺序一致
            seq = sorted(lines[yk], key=lambda t: t[1]) # 按字符的 x0 坐标排序，确保字符在行内的顺序正确
            words = [] # 存储当前行的单词列表，每个单词包含文本及其位置信息
            cur = "" # 当前正在构建的单词文本
            cx0 = None # 当前单词的起始 x 坐标
            prev_x1 = None # 上一个字符的结束 x 坐标
            line_text = ""  # 当前行的完整文本内容
            gap_sz = None   # 当前字符间距阈值，用于判断是否需要分词
            for ch, x0, x1, sz, top in seq: # 遍历当前行的字符信息
                line_text += ch # 将字符添加到当前行的完整文本内容中
                if cur == "": # 如果当前单词为空，说明这是新单词的开始
                    cur = ch # 将当前字符作为新单词的起始字符
                    cx0 = x0  # 将当前字符的起始 x 坐标作为新单词的起始坐标
                    gap_sz = max(0.5, sz * 0.30) # 设置当前字符间距阈值，确保至少为 0.5 或当前字符字号的 30%
                else: # 如果当前单词不为空，说明正在构建一个单词
                    if x0 - prev_x1 > gap_sz: # 如果当前字符的起始 x 坐标与上一个字符的结束 x 坐标之间的间距大于阈值，说明这是一个新单词
                        words.append((cur, cx0, prev_x1, top)) # 将当前单词及其位置信息添加到单词列表中
                        cur = ch  # 将当前字符作为新单词的起始字符
                        cx0 = x0 # 将当前字符的起始 x 坐标作为新单词的起始坐标
                    else:  # 如果当前字符与上一个字符之间的间距不大于阈值，说明它们属于同一个单词
                        cur += ch   # 将当前字符添加到当前单词中
                    gap_sz = max(0.5, sz * 0.30) # 更新当前字符间距阈值，确保至少为 0.5 或当前字符字号的 30%    
                prev_x1 = x1 # 更新上一个字符的结束 x 坐标为当前字符的结束 x 坐标
            if cur: # 如果当前单词不为空，说明还有一个未添加的单词
                words.append((cur, cx0, prev_x1, top)) # 将最后一个单词及其位置信息添加到单词列表中
            page_words.append((yk, words)) # 将当前行的 y 坐标及其单词列表添加到页面的行列表中
            page_text.append((yk, line_text))  # 将当前行的 y 坐标及其完整文本内容添加到页面的行文字列表中
        pages_words[pno + 1] = page_words # 将当前页的行列表存储到页面字典中，页码从 1 开始
        pages_text[pno + 1] = page_text # 将当前页的行文字列表存储到页面字典中，页码从 1 开始
        pages_words[pno + 1] = page_words # 将当前页的行列表存储到页面字典中，页码从 1 开始
        pages_text[pno + 1] = page_text # 将当前页的行文字列表存储到页面字典中，页码从 1 开始
    pdf.close() # 关闭 PDF 文件，释放资源
    return pages_words, pages_text, page_count # 返回  行列表、列表、总页数


def _get_char_boxes(pdf_path, pageno):
    try:
        with pdfplumber.open(pdf_path) as p:
            return [
                (c["x0"], c["top"], c["x1"], c["bottom"])
                for c in p.pages[pageno].chars
            ]
    except Exception:
        return None


def extract_designators(pages_words):
    """从解码文字中提取器件位号：{位号: [页码,...]}"""
    tokens = defaultdict(list)
    all_words = []
    for pno, pwords in pages_words.items():
        for y, words in pwords:
            for w, x0, x1, top in words:
                up = w.upper().strip()
                all_words.append((pno, y, up, x0, x1, top))
                if DESIGNATOR_RE.match(up) and up[0] in PREFIXES_PARTS:
                    tokens[up].append(pno)
    return tokens, all_words


def build_pdf_designator_annotations(pages_words):
    """给 PDF 里每个位号找邻近标注词（值/封装/型号）。
    返回 {位号: {'pages':[页码], 'near':[附近词,去重]}}。
    """
    tokens, _ = extract_designators(pages_words)
    # 每页的所有词
    page_words = {} # 
    for pno, pwords in pages_words.items():
        plist = []
        for y, words in pwords:
            for w, x0, x1, top in words:
                if re.match(r"^[\w.+/@-]{2,}$", w):
                    plist.append({"y": y, "x": x0, "w": w.upper()})
        page_words[pno] = plist

    out = {}
    for des, pages in tokens.items():
        des_pos = None
        for pno, pwords in pages_words.items():
            if pno not in pages:
                continue
            for y, words in pwords:
                for w, x0, x1, top in words:
                    if w.upper() == des:
                        des_pos = (pno, y, x0)
                        break
                if des_pos:
                    break
            if des_pos:
                break
        near = []
        if des_pos:
            pno, dy, dx = des_pos
            for it in page_words.get(pno, []):
                if abs(it["y"] - dy) <= 120 and abs(it["x"] - dx) <= 260 and it["w"] != des:
                    near.append(it["w"])
        out[des] = {"pages": pages, "near": sorted(set(near))[:15]}
    return out


def app_smart(v):
    """模块级包装，避免重复定义 smart_norm_value。"""
    try:
        return smart_norm_value(v)
    except Exception:
        return v


def compare_pdf_to_excel(pdf_path, bom_path, primary="Part Reference"):
    """以 PDF 位号为准，检查 Excel 中是否存在且 Value 是否疑似一致。
    返回 rows 与 stats。"""
    pages_words, pages_text, _ = decode_pdf_blocks(pdf_path)
    annotations = build_pdf_designator_annotations(pages_words)

    headers, data = load_bom(bom_path)
    primary = primary if primary in headers else headers[0]
    pidx = headers.index(primary)
    vidx = headers.index("Value") if "Value" in headers else None
    fidx = headers.index("PCB Footprint") if "PCB Footprint" in headers else None
    qidx = headers.index("Quantity") if "Quantity" in headers else None

    excel_index = {}
    for row in data:
        for d in split_designators_text(row["values"][pidx]):
            excel_index.setdefault(d.upper(), []).append(row)

    rows = []
    n_found = n_pdf_only = n_ok = n_may = 0
    for des in sorted(annotations.keys()):
        ann = annotations[des]
        near = ann["near"]
        near_text = "; ".join(near[:8])
        excel_rows = excel_index.get(des)
        if not excel_rows:
            n_pdf_only += 1
            rows.append({"item": des, "status": "PDF有Excel无", "valueA": "", "valueB": "",
                         "footA": "", "footB": "", "near": near_text, "qty": ""})
            continue
        n_found += 1
        er = excel_rows[0]
        ev = er["values"][vidx] if vidx is not None else ""
        ef = er["values"][fidx] if fidx is not None else ""
        eq = er["values"][qidx] if qidx is not None else ""
        # 值/封装是否能在邻近标注中找到（直接命中 / 归一化命中 / 短数值允许子串命中）
        nv = app_smart(ev)
        nf = app_smart(ef)
        hit = False
        candidates = [str(x) for x in [ev, nv, ef, nf] if x]
        for c in candidates:
            c = str(c).strip().lower()
            if not c:
                continue
            for w in near:
                wl = w.lower()
                if c == wl:
                    hit = True
                    break
                if app_smart(c) == app_smart(w):
                    hit = True
                    break
                # 仅当候选词是短数值(<=5字符)时才做子串匹配(值如 49.9/10K 常用)
                # 型号/封装等长词只整词匹配，避免 "13" 误命中 "TMUX1308..."
                c_short = re.fullmatch(r"[\d.]{1,6}", c)
                if c_short and len(c) <= 5 and (c in wl or wl in c):
                    hit = True
                    break
            if hit:
                break
        if hit:
            n_ok += 1
            rows.append({"item": des, "status": "一致", "valueA": ev, "valueB": "",
                         "footA": ef, "footB": "", "near": near_text, "qty": eq})
        else:
            n_may += 1
            rows.append({"item": des, "status": "待确认(值/封装在PDF上未匹配)", "valueA": ev,
                         "valueB": "", "footA": ef, "footB": "", "near": near_text, "qty": eq})

    stats = {
        "all": {
            "主键": primary,
            "PDF位号总数": len(annotations),
            "Excel中找到": n_found,
            "PDF有Excel无": n_pdf_only,
            "值疑似一致": n_ok,
            "值疑似不一致/待确认": n_may,
        },
        "bad": [{"designator": r["item"], "row": ""} for r in rows
                if r["status"] != "一致"],
        "extra": [r["item"] for r in rows if r["status"] == "PDF有Excel无"],
        "in_bom": {r["item"] for r in rows if r["status"] != "PDF有Excel无"},
        "common": [r["item"] for r in rows if r["status"] == "一致"],
        "only_a": [r["item"] for r in rows if r["status"] == "PDF有Excel无"],
        "only_b": [],
        "mode": "pdf2excel",
    }
    return rows, stats


# ---------- BOM 读取 --------------------------------
def load_bom(bom_path, sheet_contains=None):
    """读取 xlsx，返回列名列表与行数据(索引从1开始，含表头行号)。"""
    wb = openpyxl.load_workbook(bom_path, data_only=True)
    ws = wb.active
    headers = []
    data = []
    for r in range(1, ws.max_row + 1):
        row = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            row.append(v if v is not None else "")
        if r == 1:
            headers = [str(x) for x in row]
        else:
            data.append({"header_row": r, "values": row})
    return headers, data


def split_designators_text(text):
    """把单元格里的位号串（如 'R1,R2,R3' 或 'C1,C2'）拆分为列表。"""
    items = re.split(r"[,;\s\u3001]+", str(text))
    return [x.strip().upper() for x in items if x.strip()]


# ---------- 核对逻辑 --------------------------------
class CompareResult:
    def __init__(self):
        self.rows = []          # 每行结果 dict
        self.stats = {}

    @staticmethod
    def stat(rows_designators):
        n_total = 0
        n_ok = 0
        n_bad = 0
        bad_list = []
        for row in rows_designators:
            for d in row["designators"]:
                n_total += 1
                if d["found"]:
                    n_ok += 1
                else:
                    n_bad += 1
                    bad_list.append(d)
        return n_total, n_ok, n_bad, bad_list


def check_designators(headers, data, bom_col_idx, pdf_tokens):
    """按位号核对：选中列必须是 'Part Reference' 类的位号列。"""
    col_name = headers[bom_col_idx]
    qty_idx = headers.index("Quantity") if "Quantity" in headers else None
    rows = []
    for row in data:
        rn = row["header_row"]
        raw = row["values"][bom_col_idx]
        dlist = split_designators_text(raw)
        if not dlist:
            continue
        dres = []
        for d in dlist:
            found = d in pdf_tokens
            pages = pdf_tokens.get(d, []) if found else []
            dres.append({"designator": d, "found": found, "pages": pages})
        rows.append(
            {
                "header_row": rn,
                "col_name": col_name,
                "raw": raw,
                "qty": row["values"][qty_idx] if qty_idx is not None else "",
                "designators": dres,
            }
        )
    return rows


def check_value_in_pdf(headers, data, bom_col_idx, pdf_all_words, designators_only=True):
    """全文匹配：把单元格的值作为关键词在 PDF 解码文字中查找。"""
    col_name = headers[bom_col_idx]
    rows = []
    for row in data:
        rn = row["header_row"]
        raw = row["values"][bom_col_idx]
        if raw is None or str(raw).strip() == "":
            continue
        vals = split_designators_text(raw) if designators_only else [str(raw).strip()]
        found_val = []
        for v in vals:
            hits = [p for p, y, w, x0, x1, top in pdf_all_words if w.upper() == v.upper()]
            if hits:
                found_val.append(True)
            else:
                # 子串匹配（针对值里有'/'-'/空格' 等情形）
                sub = []
                for p, y, w, x0, x1, top in pdf_all_words:
                    if v.upper() in w.upper() or w.upper() in v.upper():
                        sub.append(p)
                found_val.append(bool(sub) or bool(hits))
        rows.append(
            {
                "header_row": rn,
                "col_name": col_name,
                "raw": raw,
                "designators": [
                    {
                        "designator": v,
                        "found": any(found_val),
                        "pages": [],
                    }
                    for v in vals
                ],
            }
        )
    return rows


# ============================================================
#  对比模式：Excel vs Excel / PDF vs PDF
#  思路：把每个文件归一化成"位号集合"，再算集合差异
# ============================================================
def extract_designators_from_excel(bom_path, col_name=None):
    """从 Excel BOM 中提取全部位号(Part Reference 列优先)。返回 (位号集, 位号->行号)。"""
    headers, data = load_bom(bom_path)
    col = col_name or ("Part Reference" if "Part Reference" in headers else headers[0])
    if col not in headers:
        col = headers[0]
    idx = headers.index(col)
    mapping = {}
    for row in data:
        raw = row["values"][idx]
        for d in split_designators_text(raw):
            if re.match(r"^[A-Za-z]{1,4}\d{1,4}$", d):
                mapping.setdefault(d, row["header_row"])
    return set(mapping.keys()), mapping, col


def extract_designators_from_pdf(pdf_path):
    """从 PDF 原理图中提取全部位号。返回 (位号集, 位号->页码列表)。"""
    pages_words, pages_text, _ = decode_pdf_blocks(pdf_path)
    tokens, _ = extract_designators(pages_words)
    return set(tokens.keys()), tokens


def compare_two_sets(name_a, set_a, name_b, set_b):
    """两个位号集合对比，返回 (一致, 仅A有, 仅B有) 三组。"""
    common = sorted(set_a & set_b)
    only_a = sorted(set_a - set_b)   # A有B无
    only_b = sorted(set_b - set_a)   # B有A无
    return common, only_a, only_b


# ============================================================
#  主键比对：方式A 逐位号 / 方式B 按物料分组
# ============================================================
# 单位归一化映射：0.1uF == 100nF，1K == 1000，等
_CAP_UNITS = [
    ("mf", 1000_000), ("uf", 1000_000), ("μf", 1000_000), ("µf", 1000_000),
    ("f", 1000_000_000), ("pf", 1), ("nf", 1000), ("pico", 1), ("nano", 1000),
]
_RES_UNITS = [
    ("mohm", 0.001), ("mΩ", 0.001), ("ohm", 1), ("Ω", 1), ("ω", 1),
    ("kohm", 1000), ("kω", 1000), ("k", 1000), ("mohm", 0.001),
    ("kΩ", 1000), ("kohm", 1000), ("gohm", 1000_000_000),
]


def normalize_cap_value(s):
    """把电容值归一化为 pf：0.1uF=100nF=100000pF。"""
    s = (s or "").strip().lower().replace(" ", "")
    if not s:
        return None
    s = re.sub(r"[^0-9.a-z\u00b5\u03bc]", "", s)
    if not re.search(r"\d", s):
        return s or None
    m = re.match(r"^([\d.]+)([a-z\u00b5\u03bc]*)$", s)
    if not m:
        return s
    num, unit = m.groups()
    # 数字在前、代号在后的解析
    try:
        val = float(num)
    except Exception:
        return s
    unit = unit or "f"
    mult = 1  # 基础单位 pF
    # 从大到小匹配（f 是最小根基，先处理 μ/m/n/p）
    table = {"pf": 1, "p": 1, "nf": 1000, "n": 1000, "uf": 1_000_000,
             "μf": 1_000_000, "µf": 1_000_000, "u": 1_000_000, "mf": 1_000_000,
             "f": 1_000_000_000}
    for u0, mm in sorted(table.items(), key=lambda kv: -len(kv[0])):
        if unit == u0 or unit.endswith(u0):
            mult = mm
            break
    return round(val * mult, 3)


def normalize_res_value(s):
    """把电阻值归一化为 ohm：1K=1000，2K4=2400，4R7=4.7，100R=100。"""
    s = (s or "").strip().lower().replace(" ", "")
    if not s:
        return None     # 4R7 / 2K4 / 1M5 简写形式
    if re.match(r"^\d+[rkm][\d.]+$", s):
        m = re.match(r"^(\d+)([rkm])([\d.]+)$", s)
        base = float(m.group(1) + "." + m.group(3))
        mult = {"r": 1, "k": 1000, "m": 1_000_000}[m.group(2)]
        return round(base * mult, 3)
    # 100R / 1K5 结尾形式
    if re.match(r"^[\d.]+[rkm]$", s):
        m = re.match(r"^([\d.]+)([rkm])$", s)
        mult = {"r": 1, "k": 1000, "m": 1_000_000}[m.group(2)]
        return round(float(m.group(1)) * mult, 3)
    if "ohm" in s or "Ω" in s or "ω" in s:
        m = re.match(r"^([\d.]+)\s*([a-z\u03a9\u03c9]*)$",
                     s.replace("mω", "m").replace("mΩ", "m").replace("kΩ", "k"))
        if not m:
            return s
        num, unit = m.groups()
        mult = 1
        table = {"pohm": 1e-6, "mohm": 1e-3, "ohm": 1, "Ω": 1, "ω": 1,
                 "kohm": 1000, "kω": 1000, "mω": 1_000_000, "mohm": 1e-3}
        for u0, mm in sorted(table.items(), key=lambda kv: -len(kv[0])):
            if unit.endswith(u0):
                mult = mm
                break
        try:
            return round(float(num) * mult, 3)
        except Exception:
            return s
    # "10k" "100k" "1k" 等
    m = re.match(r"^([\d.]+)\s*([km]?)$", s)
    if m:
        num, unit = m.groups()
        try:
            mult = {"": 1, "k": 1000, "m": 1_000_000}.get(unit, 1)
            return round(float(num) * mult, 3)
        except Exception:
            return s
    return s


def smart_norm_value(v):
    """通用值归一化：数字+单位换算为统一结果；否则返回清理后字符串。"""
    s = str(v or "").strip()
    if not s:
        return None
    s2 = s.lower()

    # 电容类：F / uF / nF / pF
    if any(unit in s2 for unit in ["uf", "μf", "µf", "nf", "pf", "mf"]):
        r = normalize_cap_value(s)
        if r is not None:
            return r
    # 电阻类：Ω/ohm/K(欧)/R 缩写
    if (any(unit in s2 for unit in ["ohm", "Ω", "ω", "k", "r"]) and
            "f" not in s2 and "uf" not in s2 and "µ" not in s2 and "μ" not in s2):
        r = normalize_res_value(s)
        if r is not None and r != s:
            return r
    # 纯数字：统一成数值（1 vs 1.0）
    if re.fullmatch(r"[0-9.]+", s):
        try:
            return float(s)
        except Exception:
            return s
    return s


def _px(a, b):
    """比较两个字段值（支持归一化），返回 (是否一致, 规范值a, 规范值b)。"""
    na, nb = smart_norm_value(a), smart_norm_value(b)
    if na is None and nb is None:
        return True, "", ""
    eq = (na == nb) or (str(na).upper() == str(nb).upper())
    return eq, "" if na is None else str(na), "" if nb is None else str(nb)


def _footprint_family(d):
    """把封装名归并成大类，方便比较不同写法。"""
    s = str(d or "").strip().lower().replace(" ", "")
    fam = {
        "capc1005": "0402", "capc0603": "0201", "capc1608": "0603", "capc2012": "0805",
        "capc3216": "1206", "resc1005": "0402", "resc0603": "0201", "resc1608": "0603",
        "resc2013": "0805", "indc1005": "0402", "res": "",
        "0402": "0402", "0201": "0201", "0603": "0603", "0805": "0805", "1206": "1206",
    }
    for prefix, f in fam.items():
        if s.startswith(prefix):
            return f
    return s or ""


def load_excel_table(path_a, path_b):
    """加载两个 Excel，返回 (表头A, 行A, 表头B, 行B)。"""
    ha, da = load_bom(path_a)
    hb, db = load_bom(path_b)
    return ha, da, hb, db


def compare_excel_detail(path_a, path_b, primary="Part Reference",
                         fields=("Value", "Quantity", "Manufacturer PN")):
    """方式A：逐位号比对。
    主键=primary列(默认 Part Reference)，两个文件同一位号时逐字段比较。
    返回 rows(含 'item','status','field_diffs') 与 stats。
    """
    ha, da, hb, db = load_excel_table(path_a, path_b)
    primary = primary if primary in ha else (ha[0] if ha else "Part Reference")

    def build_index(headers, data):
        idx = {}
        for row in data:
            raw = row["values"][headers.index(primary)] if primary in headers else ""
            for d in split_designators_text(raw):
                if re.match(r"^[A-Za-z]{1,4}\d{1,4}$", d):
                    idx.setdefault(d, []).append(row)
        return idx

    ia = build_index(ha, da)
    ib = build_index(hb, db)
    keys = sorted(set(ia.keys()) | set(ib.keys()))

    rows = []
    n_ok, n_diff, n_only_a, n_only_b = 0, 0, 0, 0
    for k in keys:
        rows_a = ia.get(k, [])
        rows_b = ib.get(k, [])
        if not rows_b:
            n_only_a += 1
            rows.append({"item": k, "status": "仅文件A",
                         "a": rows_a[0], "b": None, "field_diffs": [], "hdra": ha, "hdrb": hb})
            continue
        if not rows_a:
            n_only_b += 1
            rows.append({"item": k, "status": "仅文件B",
                         "a": None, "b": rows_b[0], "field_diffs": [], "hdra": ha, "hdrb": hb})
            continue
        # 同一位号：比较字段
        ra, rb = rows_a[0], rows_b[0]
        diffs = []
        for f in fields:
            fa = ha.index(f) if f in ha else None
            fb = hb.index(f) if f in hb else None
            va = ra["values"][fa] if fa is not None else None
            vb = rb["values"][fb] if fb is not None else None
            eq, na, nb = _px(va, vb)
            if not eq:
                diffs.append({"field": f, "a": "" if na is None else na,
                              "b": "" if nb is None else nb, "raw_a": va, "raw_b": vb})
        if diffs:
            n_diff += 1
            rows.append({"item": k, "status": "字段不一致", "a": ra, "b": rb,
                         "field_diffs": diffs, "hdra": ha, "hdrb": hb})
        else:
            n_ok += 1
            rows.append({"item": k, "status": "一致", "a": ra, "b": rb,
                         "field_diffs": [], "hdra": ha, "hdrb": hb})

    stats = {
        "all": {
            "主键": primary,
            "比对字段": "、".join(fields),
            "位号总数": len(keys),
            "一致": n_ok,
            "字段不一致": n_diff,
            "仅文件A有": n_only_a,
            "仅文件B有": n_only_b,
        },
        "bad": [{"designator": r["item"], "row": ""} for r in rows if r["status"] not in ("一致", "仅文件B")],
        "extra": [r["item"] for r in rows if r["status"] == "仅文件B"],
        "in_bom": {r["item"] for r in rows if r["status"] != "仅文件B"},
        "common": [r["item"] for r in rows if r["status"] in ("一致", "字段不一致")],
        "only_a": [r["item"] for r in rows if r["status"] == "仅文件A"],
        "only_b": [r["item"] for r in rows if r["status"] == "仅文件B"],
        "mode": "detail",
    }
    return rows, stats


def compare_excel_group(path_a, path_b, group_fields=("Value", "Manufacturer PN"),
                        qty_col="Quantity", ref_col="Part Reference"):
    """方式B：按物料分组比对。
    以 group_fields 为键分组，比较每组位号集合与数量。
    """
    ha, da, hb, db = load_excel_table(path_a, path_b)

    def group_index(headers, data):
        g = {}
        for row in data:
            vals = row["values"]
            key = tuple((smart_norm_value(vals[headers.index(f)]) if f in headers
                         else None) for f in group_fields)
            # 料号空值：把 None 视为能匹配
            key = tuple("" if k is None else str(k) for k in key)
            refs = split_designators_text(vals[headers.index(ref_col)])
            g.setdefault(key, []).extend(refs)
        return g

    ga = group_index(ha, da)
    gb = group_index(hb, db)
    all_keys = sorted(set(ga.keys()) | set(gb.keys()))

    rows = []
    n_ok, n_diff, n_only_a, n_only_b = 0, 0, 0, 0
    for key in all_keys:
        ra = sorted(ga.get(key, []))
        rb = sorted(gb.get(key, []))
        if not rb:
            n_only_a += 1
            rows.append({"item": " | ".join(str(k) for k in key), "status": "仅文件A",
                         "refs_a": ra, "refs_b": [], "qty_a": len(ra), "qty_b": 0})
            continue
        if not ra:
            n_only_b += 1
            rows.append({"item": " | ".join(str(k) for k in key), "status": "仅文件B",
                         "refs_a": [], "refs_b": rb, "qty_a": 0, "qty_b": len(rb)})
            continue
        # 同组：比较位号集合
        only_in_a = sorted(set(ra) - set(rb))
        only_in_b = sorted(set(rb) - set(ra))
        if only_in_a or only_in_b:
            n_diff += 1
            rows.append({"item": " | ".join(str(k) for k in key), "status": "位号集合差异",
                         "refs_a": ra, "refs_b": rb, "qty_a": len(ra), "qty_b": len(rb),
                         "only_a": only_in_a, "only_b": only_in_b})
        else:
            n_ok += 1
            rows.append({"item": " | ".join(str(k) for k in key), "status": "一致",
                         "refs_a": ra, "refs_b": rb, "qty_a": len(ra), "qty_b": len(rb)})

    stats = {
        "all": {
            "分组键": "、".join(group_fields),
            "物料分组数": len(all_keys),
            "一致": n_ok,
            "位号集合差异": n_diff,
            "仅文件A有": n_only_a,
            "仅文件B有": n_only_b,
        },
        "bad": [{"designator": r["item"], "row": ""} for r in rows
                if r["status"] in ("位号集合差异", "仅文件A")],
        "extra": [r["item"] for r in rows if r["status"] == "仅文件B"],
        "in_bom": {r["item"] for r in rows if r["status"] != "仅文件B"},
        "common": [r["item"] for r in rows if r["status"] == "一致"],
        "only_a": [r["item"] for r in rows if r["status"] == "仅文件A"],
        "only_b": [r["item"] for r in rows if r["status"] == "仅文件B"],
        "mode": "group",
    }
    return rows, stats


def compare_excel_excel(path_a, path_b, col_name):
    """两个 Excel BOM 位号对比。"""
    set_a, map_a, col_a = extract_designators_from_excel(path_a, col_name)
    set_b, map_b, col_b = extract_designators_from_excel(path_b, col_name)
    common, only_a, only_b = compare_two_sets(os.path.basename(path_a), set_a,
                                              os.path.basename(path_b), set_b)
    rows = []
    for d in common:
        rows.append({"item": d, "status": "一致", "a": map_a.get(d, ""), "b": map_b.get(d, ""),
                     "a_pages": "", "b_pages": ""})
    for d in only_a:
        rows.append({"item": d, "status": "仅文件A", "a": map_a.get(d, ""), "b": "",
                     "a_pages": "", "b_pages": ""})
    for d in only_b:
        rows.append({"item": d, "status": "仅文件B", "a": "", "b": map_b.get(d, ""),
                     "a_pages": "", "b_pages": ""})
    stats = {
        "all": {
            "文件A位号数": len(set_a),
            "文件B位号数": len(set_b),
            "两文件一致": len(common),
            "仅文件A有": len(only_a),
            "仅文件B有": len(only_b),
            "对比列": col_a,
        },
        "bad": [{"designator": d, "row": map_a.get(d, "")} for d in only_a],
        "extra": only_b,
        "in_bom": set_a,
        "common": common,
        "only_a": only_a,
        "only_b": only_b,
    }
    return rows, stats


def compare_pdf_pdf(path_a, path_b):
    """两个 PDF 原理图位号对比。"""
    set_a, tokens_a = extract_designators_from_pdf(path_a)
    set_b, tokens_b = extract_designators_from_pdf(path_b)
    common, only_a, only_b = compare_two_sets(os.path.basename(path_a), set_a,
                                              os.path.basename(path_b), set_b)
    rows = []
    for d in common:
        rows.append({"item": d, "status": "一致",
                     "a": "", "b": "",
                     "a_pages": ",".join(str(p) for p in tokens_a.get(d, [])),
                     "b_pages": ",".join(str(p) for p in tokens_b.get(d, []))})
    for d in only_a:
        rows.append({"item": d, "status": "仅文件A", "a": "", "b": "",
                     "a_pages": ",".join(str(p) for p in tokens_a.get(d, [])), "b_pages": ""})
    for d in only_b:
        rows.append({"item": d, "status": "仅文件B", "a": "", "b": "",
                     "a_pages": "", "b_pages": ",".join(str(p) for p in tokens_b.get(d, []))})
    stats = {
        "all": {
            "PDF A 位号数": len(set_a),
            "PDF B 位号数": len(set_b),
            "两文件一致": len(common),
            "仅 PDF A 有": len(only_a),
            "仅 PDF B 有": len(only_b),
        },
        "bad": [{"designator": d, "row": ""} for d in only_a],
        "extra": only_b,
        "in_bom": set_a,
        "common": common,
        "only_a": only_a,
        "only_b": only_b,
    }
    return rows, stats


# ---------- 报告生成 --------------------------------
def gen_report(bom_path, rows, stats, pdf_tokens, pages_text, out_dir):
    ts = time.strftime("%Y%m%d_%H%M%S")
    # 根据 row 结构自动识别模式：对比模式 rows 含 "item"/"status" 且无 "designators"
    is_compare = bool(rows) and "designators" not in rows[0]
    if is_compare:
        xlsx_path = os.path.join(out_dir, f"对比报告_{ts}.xlsx")
        txt_path = os.path.join(out_dir, f"对比报告_{ts}.txt")
    else:
        xlsx_path = os.path.join(out_dir, f"BOM核对报告_{ts}.xlsx")
        txt_path = os.path.join(out_dir, f"BOM核对报告_{ts}.txt")

    wb = openpyxl.Workbook()
    ws = wb.active
    if is_compare:
        mode = stats.get("mode", "set")
        if mode == "pdf2excel":
            ws.title = "PDF→Excel 器件核对"
            ws.append(["PDF位号", "核对结果", "Excel值(Value)", "Excel封装(Footprint)",
                       "Excel数量", "PDF附近标注"])
            for row in rows:
                ws.append([row["item"], row["status"], row["valueA"], row["footA"],
                           row["qty"], row["near"]])
        elif mode == "detail":
            ws.title = "逐位号对比"
            ws.append(["位号(主键)", "结果", "字段", "文件A值", "文件B值"])
            for row in rows:
                if row["field_diffs"]:
                    for d in row["field_diffs"]:
                        ws.append([row["item"], row["status"], d["field"], d["a"], d["b"]])
                else:
                    ws.append([row["item"], row["status"], "", "", ""])
        elif mode == "group":
            ws.title = "按物料分组"
            ws.append(["物料分组(键)", "结果", "数量A", "数量B", "位号A", "位号B"])
            for row in rows:
                ws.append([row["item"], row["status"], row["qty_a"], row["qty_b"],
                           ", ".join(row.get("refs_a", [])), ", ".join(row.get("refs_b", []))])
        else:
            ws.title = "对比明细"
            ws.append(["器件/位号", "结果", "文件A行号/页码", "文件B行号/页码"])
            for row in rows:
                ws.append([row["item"], row["status"], row["a_pages"] or row["a"],
                           row["b_pages"] or row["b"]])
    else:
        ws.title = "核对明细"
        ws.append(
            ["BOM行号", "列", "原值", "位号/值", "PDF中是否存在", "所在页码", "结果"]
        )
        for row in rows:
            for d in row["designators"]:
                ws.append(
                    [
                        row["header_row"],
                        row["col_name"],
                        row["raw"],
                        d["designator"],
                        "是" if d["found"] else "否",
                        ",".join(str(p) for p in d["pages"]) if d["found"] else "",
                        "一致" if d["found"] else "不一致",
                    ]
                )

    # Sheet2: 汇总
    ws2 = wb.create_sheet("汇总")
    ws2.append(["统计项", "数量"])
    for k, v in stats["all"].items():
        ws2.append([k, v])
    ws2.append([])
    if is_compare:
        ws2.append(["仅文件A有：", len(stats["bad"])])
        for b in stats["bad"]:
            ws2.append(["", b["designator"]])
        ws2.append(["仅文件B有：", len(stats["extra"])])
        for e in stats["extra"]:
            ws2.append(["", e])
    else:
        ws2.append(["BOM未在PDF中找到的位号：", len(stats["bad"])])
        for b in stats["bad"]:
            ws2.append(["", b["designator"], "", "不一致"])
        # Sheet3: PDF 中出现的所有器件位号
        ws3 = wb.create_sheet("PDF全部位号")
        ws3.append(["位号", "出现页码", "是否在BOM"])
        in_bom = stats.get("in_bom", set())
        for t in sorted(pdf_tokens.keys(),
                        key=lambda s: (s[0], int(re.search(r"\d+", s).group()) if re.search(r"\d+", s) else 0)):
            ws3.append([t, ",".join(str(p) for p in pdf_tokens[t]), "是" if t in in_bom else "否"])
    wb.save(xlsx_path)

    # txt 报告
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write("=" * 70 + "\n")
        if is_compare:
            f.write("文件对比报告（Excel vs Excel / PDF vs PDF / PDF→Excel）\n")
        else:
            f.write("BOM 与 PDF 原理图 器件核对报告\n")
        f.write("来源: %s\n" % bom_path)
        f.write("时间: %s\n" % time.strftime("%Y-%m-%d %H:%M:%S"))
        f.write("=" * 70 + "\n\n")
        f.write("【汇总】\n")
        for k, v in stats["all"].items():
            f.write("  %-30s: %s\n" % (k, v))
        if is_compare:
            mode = stats.get("mode", "set")
            if mode == "pdf2excel":
                f.write("\n【疑似不一致 / 待确认】\n")
                for b in stats["bad"]:
                    f.write("  %s\n" % b["designator"])
                f.write("\n【PDF有Excel无】\n")
                for t in sorted(stats["extra"]):
                    f.write("  %s\n" % t)
                f.write("\n【明细】\n")
                for row in rows:
                    f.write("  [%s] %s  Excel值=%s 封装=%s  附近标注=%s\n"
                            % (row["status"], row["item"], row["valueA"], row["footA"], row["near"]))
            else:
                f.write("\n【仅文件A有】\n")
                for b in stats["bad"]:
                    f.write("  %s\n" % b["designator"])
                f.write("\n【仅文件B有】\n")
                for t in sorted(stats["extra"]):
                    f.write("  %s\n" % t)
                f.write("\n【明细】\n")
                if mode == "detail":
                    for row in rows:
                        if row["field_diffs"]:
                            f.write("  [%s] %s\n" % (row["status"], row["item"]))
                            for d in row["field_diffs"]:
                                f.write("       字段 %-20s  A=%s  B=%s\n"
                                        % (d["field"], d["a"] or "(空)", d["b"] or "(空)"))
                        else:
                            f.write("  [%s] %s\n" % (row["status"], row["item"]))
                elif mode == "group":
                    for row in rows:
                        f.write("  [%s] %s  A数量=%d B数量=%d\n"
                                % (row["status"], row["item"], row["qty_a"], row["qty_b"]))
                        if row.get("only_a"):
                            f.write("       A独有位号: %s\n" % ", ".join(row["only_a"]))
                        if row.get("only_b"):
                            f.write("       B独有位号: %s\n" % ", ".join(row["only_b"]))
                else:
                    for row in rows:
                        f.write("  %-6s | %-10s | A:%s | B:%s\n"
                                % (row["status"], row["item"], row["a_pages"] or row["a"],
                                   row["b_pages"] or row["b"]))
        else:
            f.write("\n【不一致 - BOM位号在PDF中未找到】\n")
            for b in stats["bad"]:
                f.write("  %s  (BOM第%d行)\n" % (b["designator"], b["row"]))
            f.write("\n【PDF中出现但不在BOM的位号】\n")
            for t in sorted(stats["extra"]):
                f.write("  %s  页码:%s\n" % (t, ",".join(str(p) for p in pdf_tokens[t])))
            f.write("\n【明细】\n")
            for row in rows:
                for d in row["designators"]:
                    mark = "OK " if d["found"] else "MISS"
                    f.write(
                        "%s | 行%-3d | %-6s | %-10s | %s\n"
                        % (mark, row["header_row"], row["col_name"], d["designator"], row["raw"])
                    )
    return xlsx_path, txt_path


# ============================================================
#  美化 GUI（含二级子窗口）
# ============================================================
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# ---------- 配色主题 --------------------------------
class Theme:
    BG        = "#F4F6FB"   # 主背景色
    CARD      = "#FFFFFF"   # 卡片底色
    BRAND     = "#2C6FBB"   # 顶部横幅+按钮 主色调（改成你喜欢的主色）
    BRAND_DK  = "#1E4E84"   #   主色加深
    BRAND_LT  = "#E8F0FA"   #   主色浅底（表头用）
    TEXT      = "#23303B"   # 正文文字色
    SUBTEXT   = "#6B7A8D"   # 次要文字色
    LINE      = "#DFE6F0"   # 边框/分隔线色
    OK_BG     = "#E7F7EA"   # 一致行的绿底
    OK_FG     = "#1E7B34"   # 一致行的绿字
    BAD_BG    = "#FDEAEA"   # 不一致行的红底
    BAD_FG    = "#C0392B"   # 不一致行的红字
    WARN_BG   = "#FFF4E0"   # 警告行的黄底
    WARN_FG   = "#996A00"   # 警告行的黄字
    BUTTON_OK = "#2E8B57"


# ---------- Windows 高DPI 适配 ----------
# 解决 tkinter 在 Windows 高分屏下文字模糊、发虚的问题。
def enable_dpi_awareness():
    """在创建任何窗口前调用，让程序感知 Windows 缩放比例(DPI)。"""
    if sys.platform == "win32":
        try:
            import ctypes
            # 0=无效, 1=系统DPI感知, 2=每显示器DPI感知(最清晰)
            ctypes.windll.shcore.SetProcessDpiAwareness(2)
        except Exception:
            try:
                import ctypes
                ctypes.windll.user32.SetProcessDPIAware()
            except Exception:
                pass


def detect_scaling(root):
    """根据屏幕 DPI 计算字体缩放系数，返回 (scale, 单位字体尺寸)。"""
    try:
        import ctypes
        dpi = float(ctypes.windll.shcore.GetScaleFactorForDevice(0))
    except Exception:
        dpi = 96.0
    scale = dpi / 96.0
    # 用 tkinter 的 scaling 保持字体物理尺寸一致
    try:
        root.tk.call("tk", "scaling", scale)
    except Exception:
        pass
    return scale


# 统一字体规格(避免大小/风格不一致)
def _f(*size_weight):
    """生成统一字体元组，默认用 Segoe UI（Windows 自带现代字体，清晰圆润）。
    用法: _f(10)   → 常规10号;   _f(10, "bold") → 粗体10号
    """
    size = size_weight[0]
    weight = size_weight[1] if len(size_weight) > 1 else "normal"
    return ("Segoe UI", int(size), weight)


def _style(theme, scale=1.0):
    s = ttk.Style()
    try:
        s.theme_use("clam")
    except Exception:
        pass
    base = max(10, round(10 * scale))
    s.configure(".", background=theme.BG, foreground=theme.TEXT, font=_f(base))
    s.configure("Card.TFrame", background=theme.CARD)
    s.configure("Card.TLabelframe", background=theme.CARD, bordercolor=theme.LINE,
                relief="solid", borderwidth=1)
    s.configure("Card.TLabelframe.Label", background=theme.CARD, foreground=theme.BRAND_DK,
                font=_f(base, "bold"))
    s.configure("TLabel", background=theme.BG, foreground=theme.TEXT, font=_f(base))
    s.configure("CLabel.TLabel", background=theme.CARD, font=_f(base))
    s.configure("Sub.TLabel", background=theme.CARD, foreground=theme.SUBTEXT,
                font=_f(max(9, base - 1)))
    s.configure("Title.TLabel", background=theme.CARD, foreground=theme.BRAND_DK,
                font=_f(max(15, round(base * 1.5)), "bold"))
    s.configure("Count.TLabel", background=theme.CARD, font=_f(max(11, base), "bold"))
    s.configure("OK.TLabel", background=theme.CARD, foreground=theme.OK_FG, font=_f(base, "bold"))
    s.configure("BAD.TLabel", background=theme.CARD, foreground=theme.BAD_FG, font=_f(base, "bold"))

    s.configure("TButton", background="#E7EDF6", foreground=theme.TEXT, bordercolor=theme.LINE,
                padding=(12, 6), font=_f(base))
    s.map("TButton", background=[("active", "#D5E2F4")])
    s.configure("Primary.TButton", background=theme.BRAND, foreground="#FFFFFF", bordercolor=theme.BRAND,
                padding=(16, 8), font=_f(base, "bold"))
    s.map("Primary.TButton", background=[("active", theme.BRAND_DK)])
    s.configure("Danger.TButton", background="#E8B4B4", foreground="#8B0000", bordercolor="#C88080",
                padding=(14, 6), font=_f(base))
    s.map("Danger.TButton", background=[("active", "#DCA0A0")])
    s.configure("Success.TButton", background="#5B9E6B", foreground="#FFF", bordercolor="#4A8B58",
                padding=(14, 6), font=_f(base))
    s.map("Success.TButton", background=[("active", "#478457")])

    s.configure("TEntry", fieldbackground="#FFF", foreground=theme.TEXT, bordercolor=theme.LINE,
                font=_f(base))
    s.configure("TCombobox", fieldbackground="#FFF", foreground=theme.TEXT,
                bordercolor=theme.LINE, arrowcolor=theme.BRAND, font=_f(base))

    s.configure("Treeview", background="#FFFFFF", fieldbackground="#FFFFFF", foreground=theme.TEXT,
                rowheight=max(26, round(26 * scale)), bordercolor=theme.LINE, font=_f(base))
    s.configure("Treeview.Heading", background=theme.BRAND_LT, foreground=theme.BRAND_DK,
                font=_f(base, "bold"), relief="flat")
    s.map("Treeview", background=[("selected", "#BFD8F2")])
    s.configure("Vertical.TScrollbar", background="#DAE4F0", troughcolor=theme.BG, arrowcolor=theme.BRAND_DK)
    s.configure("Horizontal.TScrollbar", background="#DAE4F0", troughcolor=theme.BG, arrowcolor=theme.BRAND_DK)


def _make_card(parent, title, padding=(14, 12, 14, 12)):
    lf = ttk.LabelFrame(parent, text=title, style="Card.TLabelframe", padding=padding)
    return lf


def _icon_text(label, color, bg):
    w = tk.Label(label.master, text=" ", bg=bg, fg=color, padx=4)
    return w


# ---------- 二级窗口：PDF 文字查看 --------------------
def open_pdf_text_win(root, app):
    from tkinter import Text
    win = tk.Toplevel(root)
    win.title("PDF 原理图解码文字")
    win.geometry("760x560")
    win.configure(bg=Theme.BG)
    win.transient(root)
    win.grab_set()

    head = ttk.Frame(win, style="Card.TFrame", padding=10)
    head.pack(fill="x", padx=10, pady=(10, 6))
    ttk.Label(head, text="PDF 每页解码出的文字（含位置坐标）",
              style="Title.TLabel").pack(side="left")
    ttk.Label(head, text="", style="Sub.TLabel").pack(side="left")

    ttk.Style().configure("page.Tab", padding=(20, 6))

    pw = ttk.Panedwindow(win, orient="horizontal")
    pw.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    left = ttk.Frame(pw)
    pw.add(left, weight=0)
    ttk.Label(left, text="选择页码:", style="Sub.TLabel").pack(pady=(4, 2))
    lb = tk.Listbox(left, width=10, height=28, activestyle="dotbox",
                    selectbackground=Theme.BRAND, selectforeground="#fff",
                    font=_f(10), bg="#FFF", fg=Theme.TEXT, relief="flat",
                    highlightbackground=Theme.LINE, highlightthickness=1)
    lb.pack(fill="y", expand=True)

    right = ttk.Frame(pw)
    pw.add(right, weight=1)
    txt = Text(right, wrap="none", font=_f(10),
               bg="#FFFFFF", fg=Theme.TEXT, insertbackground=Theme.BRAND,
               relief="flat", padx=10, pady=8)
    vs = ttk.Scrollbar(right, orient="vertical", command=txt.yview)
    hs = ttk.Scrollbar(right, orient="horizontal", command=txt.xview)
    txt.configure(yscrollcommand=vs.set, xscrollcommand=hs.set)
    txt.grid(row=0, column=0, sticky="nsew")
    vs.grid(row=0, column=1, sticky="ns")
    hs.grid(row=1, column=0, sticky="ew")
    right.rowconfigure(0, weight=1)
    right.columnconfigure(0, weight=1)

    pages_store = {}
    for pno in sorted(app.get("pdf_pages_text", {})):
        lb.insert("end", f"第 {pno} 页")
        pl = app["pdf_pages_text"][pno]
        body = []
        for y, line in pl:
            body.append(f"{y:<8.1f}  {line}")
        pages_store[pno] = "\n".join(body)

    def show(_=None):
        sel = lb.curselection()
        if not sel:
            return
        pno = int(lb.get(sel[0]).replace("第 ", "").replace(" 页", ""))
        txt.delete("1.0", "end")
        txt.insert("1.0", pages_store.get(pno, "(无文字)"))

    lb.bind("<<ListboxSelect>>", show)
    if lb.size():
        lb.selection_set(0)
        show()


# ---------- 二级窗口：位号详情 ------------------------
def open_detail_win(root, app, designator, pages, row, col, raw):
    win = tk.Toplevel(root)
    win.title(f"位号详情  {designator}")
    win.geometry("640x420")
    win.configure(bg=Theme.BG)
    win.transient(root)
    win.grab_set()

    card = _make_card(win, "查询信息")
    card.pack(fill="x", padx=12, pady=10)
    grid = ttk.Frame(card, style="Card.TFrame")
    grid.pack(fill="x")
    rows = [
        ("位号", designator),
        ("BOM行", str(row)),
        ("核对列", col),
        ("原值(BOM)", str(raw)),
        ("PDF页码", ", ".join(str(p) for p in pages) if pages else "未找到"),
        ("结果", "一致" if pages else "不一致"),
    ]
    for i, (k, v) in enumerate(rows):
        ttk.Label(grid, text=k, style="Sub.TLabel", width=10).grid(row=i, column=0, sticky="ne", pady=2)
        ttk.Label(grid, text=v, style="CLabel.TLabel", anchor="w").grid(row=i, column=1, sticky="w", padx=8, pady=2)
        if k == "结果":
            if pages:
                ttk.Label(grid, text="一致", style="OK.TLabel").grid(row=i, column=2, sticky="w")
            else:
                ttk.Label(grid, text="不一致", style="BAD.TLabel").grid(row=i, column=2, sticky="w")

    ctx = _make_card(win, "位号在 PDF 中出现的行（上下文）")
    ctx.pack(fill="both", expand=True, padx=12, pady=(0, 12))
    area = tk.Text(ctx, height=12, font=_f(10), bg="#FFF",
                   fg=Theme.TEXT, relief="flat", padx=8, pady=6)
    area.pack(side="left", fill="both", expand=True)
    sv = ttk.Scrollbar(ctx, orient="vertical", command=area.yview)
    sv.pack(side="right", fill="y")
    area.configure(yscrollcommand=sv.set)
    area.insert("1.0", "（未找到该位号在 PDF 中的上下文）" if not pages else "")
    if pages:
        # 显示每页包含该位号的行
        de = app.get("designator_context", {})
        lines = de.get(designator, [])
        area.delete("1.0", "end")
        if lines:
            for ln in lines:
                area.insert("end", ln + "\n")
        else:
            area.insert("1.0", "该位号已找到（页码见上）。可打开“PDF文字”窗口查看上下文。")


# ---------- 二级窗口：不一致清单 ----------------------
def open_mismatch_win(root, app):
    stats = app.get("stats")
    if not stats:
        messagebox.showinfo("提示", "请先执行“开始核对”。")
        return
    win = tk.Toplevel(root)
    win.title("不一致清单")
    win.geometry("760x560")
    win.configure(bg=Theme.BG)
    win.transient(root)

    card = _make_card(win, "BOM 有、PDF 无（不一致）")
    card.pack(fill="both", expand=True, padx=12, pady=8)
    tree = ttk.Treeview(card, columns=("des", "row", "raw"), show="headings", height=12)
    tree.heading("des", text="位号")
    tree.heading("row", text="BOM 行")
    tree.heading("raw", text="原值")
    tree.column("des", width=140)
    tree.column("row", width=80)
    tree.column("raw", width=420)
    vs = ttk.Scrollbar(card, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vs.set)
    tree.pack(side="left", fill="both", expand=True)
    vs.pack(side="right", fill="y")
    for b in stats["bad"]:
        tree.insert("", "end", values=(b["designator"], b.get("row", ""), ""), tags="bad")
    tree.tag_configure("bad", background=Theme.BAD_BG, foreground=Theme.BAD_FG)

    card2 = _make_card(win, "PDF 有、BOM 无（额外器件/网络）")
    card2.pack(fill="both", expand=True, padx=12, pady=(0, 12))
    tree2 = ttk.Treeview(card2, columns=("des", "pages"), show="headings", height=10)
    tree2.heading("des", text="位号")
    tree2.heading("pages", text="PDF 页码")
    tree2.column("des", width=140)
    tree2.column("pages", width=160)
    vs2 = ttk.Scrollbar(card2, orient="vertical", command=tree2.yview)
    tree2.configure(yscrollcommand=vs2.set)
    tree2.pack(side="left", fill="both", expand=True)
    vs2.pack(side="right", fill="y")
    for e in stats.get("extra", []):
        tree2.insert("", "end", values=(e, ",".join(str(p) for p in app.get("pdf_tokens", {}).get(e, []))),
                     tags="warn")
    tree2.tag_configure("warn", background=Theme.WARN_BG, foreground=Theme.WARN_FG)


# ---------- 二级窗口：汇总统计 ------------------------
def open_summary_win(root, app):
    stats = app.get("stats")
    if not stats:
        messagebox.showinfo("提示", "请先执行“开始核对”。")
        return
    win = tk.Toplevel(root)
    win.title("核对汇总")
    win.geometry("460x300")
    win.configure(bg=Theme.BG)
    win.transient(root)
    card = _make_card(win, "统计")
    card.pack(fill="both", expand=True, padx=12, pady=12)
    for k, v in stats["all"].items():
        row = ttk.Frame(card, style="Card.TFrame")
        row.pack(fill="x", pady=3)
        ttk.Label(row, text=k, style="CLabel.TLabel", width=32, anchor="w").pack(side="left")
        ttk.Label(row, text=str(v), style="Count.TLabel", anchor="e").pack(side="right")


# ============================================================
#  主窗口
# ============================================================
def run_gui():
    import tkinter as tk
    from tkinter import ttk

    enable_dpi_awareness()          # ★ 必须在创建窗口前调用：解决模糊

    root = tk.Tk()
    root.title("BOM ↔ PDF 原理图 器件核对工具")
    scale = detect_scaling(root)    # 依据屏幕 DPI 计算缩放
    _style(Theme, scale)            # 字体随缩放联动，避免大小不一致

    W, H = 1180, 820
    if scale > 1.0:
        W, H = int(W * scale), int(H * scale)
    root.geometry(f"{W}x{H}")
    root.minsize(int(980 * scale) if scale > 1 else 980,
                 int(700 * scale) if scale > 1 else 700)
    root.configure(bg=Theme.BG, padx=8, pady=8)

    # 全局数据
    app = {
        "headers": [], "data": [], "bom_path": "", "pdf_path": "",
        "pdf_tokens": {}, "pdf_all_words": [], "pdf_pages_text": {},
        "result_rows": [], "stats": None, "designator_context": {},
        "mode": "BOM vs PDF",           # 当前对比类型
        "file_b": "",                   # 第二个文件（对比用）
        "headers_b": [], "data_b": [],
    }
    lb_xlsx = tk.StringVar()
    lb_pdf = tk.StringVar()
    lb_file_b = tk.StringVar()
    lb_compare = tk.StringVar(value="BOM vs PDF")
    lb_col = tk.StringVar()
    lb_mode = tk.StringVar(value="位号核对(推荐)")
    var_status = tk.StringVar(value="就绪：请选择对比类型与文件")

    # ---------- 顶部标题栏 ----------
    header = tk.Frame(root, bg=Theme.BRAND)
    header.pack(fill="x", pady=(0, 8))
    tk.Label(header, text="BOM ↔ PDF 器件核对与对比工具", bg=Theme.BRAND, fg="#FFFFFF",
             font=_f(max(15, round(15 * scale)), "bold")).pack(side="left", padx=14, pady=10)
    tk.Label(header, text="Excel BOM 与 PDF 原理图一致性校验 · 支持双 Excel / 双 PDF 对比",
             bg=Theme.BRAND, fg="#DCE8FB", font=_f(max(9, round(9 * scale)))).pack(side="left", padx=6)

    # ---------- 主内容区 ----------
    main = ttk.Frame(root)
    main.pack(fill="both", expand=True)

    # 左栏：设置卡片
    left_col = ttk.Frame(main)
    left_col.pack(side="left", fill="y", padx=(0, 8))
    card_compare = _make_card(left_col, "1 · 对比类型")
    card_compare.pack(fill="x", pady=(0, 8))
    _compare_type_row(card_compare, lb_compare, lb_mode, lb_col, lb_xlsx, lb_pdf, lb_file_b,
                      app, var_status)

    card_file = _make_card(left_col, "2 · 选择文件")
    card_file.pack(fill="x", pady=(0, 8))
    _file_rows(card_file, lb_compare, lb_xlsx, lb_pdf, lb_file_b, app, var_status, lb_col, root)

    card_key = _make_card(left_col, "3 · 核对设置")
    card_key.pack(fill="x", pady=(0, 8))
    _setting_rows(card_key, lb_col, lb_mode, app, lb_xlsx, lb_pdf, lb_file_b, lb_compare,
                  var_status, root)

    card_prob = _make_card(left_col, "4 · 结果窗口（二级）")
    card_prob.pack(fill="x")
    _subwin_rows(card_prob, root, app)

    # 右栏：预览与结果
    right_col = ttk.Frame(main)
    right_col.pack(side="right", fill="both", expand=True)

    _preview_card(right_col, app, lb_col)
    _result_card(right_col, root, app)

    # ---------- 底部状态栏 ----------
    status_bar = tk.Frame(root, bg=Theme.CARD, highlightbackground=Theme.LINE, highlightthickness=1)
    status_bar.pack(fill="x", pady=(8, 0))
    tk.Label(status_bar, textvariable=var_status, bg=Theme.CARD, fg=Theme.SUBTEXT,
             font=_f(max(9, round(9 * scale))), anchor="w", padx=10, pady=6).pack(fill="x")

    # ---------- 自动载入同目录文件 ----------
    root.after(200, lambda: _auto_load(app, lb_xlsx, lb_pdf, lb_col))
    root.mainloop()


# ---------- 对比类型行 ----------
def _compare_type_row(parent, lb_compare, lb_mode, lb_col, lb_xlsx, lb_pdf, lb_file_b,
                      app, var_status):
    row = ttk.Frame(parent, style="Card.TFrame")
    row.pack(fill="x", pady=3)
    ttk.Label(row, text="对比类型", style="Sub.TLabel", width=11, anchor="w").pack(side="left")
    cmb = ttk.Combobox(row, textvariable=lb_compare, width=26, state="readonly")
    cmb.pack(side="left", fill="x", expand=True)
    cmb["values"] = ["BOM vs PDF", "Excel vs Excel", "PDF vs PDF", "PDF → Excel(器件核对)"]
    cmb.current(0)
    cmb.bind("<<ComboboxSelected>>", lambda e: _on_compare_change(
        lb_compare, lb_mode, lb_col, lb_xlsx, lb_pdf, lb_file_b, app, var_status))


def _on_compare_change(lb_compare, lb_mode, lb_col, lb_xlsx, lb_pdf, lb_file_b, app, var_status):
    t = lb_compare.get()
    app["mode"] = t
    if t == "Excel vs Excel":
        lb_mode.set("位号对比")
        var_status.set("请选择两个 Excel 文件（A=基准，B=对比）")
    elif t == "PDF vs PDF":
        lb_mode.set("位号对比")
        var_status.set("请选择两个 PDF 文件（A=基准，B=对比）")
    elif t == "PDF → Excel(器件核对)":
        lb_mode.set("器件核对")
        var_status.set("以 PDF 位号为准，核对 Excel 中是否存在该器件及 Value/Footprint")
    else:
        lb_mode.set("位号核对(推荐)")
        var_status.set("请选择 BOM 与 PDF 文件")


# ---------- 文件行（三模式动态） ----------
def _file_rows(parent, lb_compare, lb_xlsx, lb_pdf, lb_file_b, app, var_status, lb_col, root):
    # 行 A（基准文件）：BOM vs PDF 时=Excel；Excel vs Excel 时=Excel A；PDF vs PDF 时=PDF A
    rA = ttk.Frame(parent, style="Card.TFrame")
    rA.pack(fill="x", pady=3)
    lab_a = ttk.Label(rA, text="文件 A", style="Sub.TLabel", width=11, anchor="w")
    lab_a.pack(side="left")
    eA = ttk.Entry(rA, textvariable=lb_xlsx, width=30)
    eA.pack(side="left", fill="x", expand=True)
    ttk.Button(rA, text="浏览", width=6,
               command=lambda: _browse_file(app, "A", lb_compare, lb_xlsx, lb_pdf, lb_file_b,
                                            var_status, lb_col, root)).pack(side="left", padx=(4, 0))

    def _bind_a():
        if lb_compare.get() in ("PDF vs PDF", "PDF → Excel(器件核对)"):
            eA.configure(textvariable=lb_pdf)
        else:
            eA.configure(textvariable=lb_xlsx)

    # 行 B（第二文件）：BOM vs PDF 时=PDF；对比模式=文件B；PDF→Excel时=Excel
    rB = ttk.Frame(parent, style="Card.TFrame")
    rB.pack(fill="x", pady=3)
    lab_b = ttk.Label(rB, text="文件 B", style="Sub.TLabel", width=11, anchor="w")
    lab_b.pack(side="left")
    eB = ttk.Entry(rB, textvariable=lb_pdf, width=30)
    eB.pack(side="left", fill="x", expand=True)
    ttk.Button(rB, text="浏览", width=6,
               command=lambda: _browse_file(app, "B", lb_compare, lb_xlsx, lb_pdf, lb_file_b,
                                            var_status, lb_col, root)).pack(side="left", padx=(4, 0))

    # PDF→Excel 模式：行B 显示 Excel，应绑定 lb_xlsx；切换时改绑定
    def _bind_b():
        if lb_compare.get() == "PDF → Excel(器件核对)":
            eB.configure(textvariable=lb_xlsx)
        else:
            eB.configure(textvariable=lb_pdf)

    # 行 C（仅对比模式的第二文件）：在对比模式下显示
    rC = ttk.Frame(parent, style="Card.TFrame")
    rC.pack(fill="x", pady=3)
    lab_c = ttk.Label(rC, text="文件 C (B)", style="Sub.TLabel", width=11, anchor="w")
    lab_c.pack(side="left")
    eC = ttk.Entry(rC, textvariable=lb_file_b, width=30)
    eC.pack(side="left", fill="x", expand=True)
    ttk.Button(rC, text="浏览", width=6,
               command=lambda: _browse_file(app, "C", lb_compare, lb_xlsx, lb_pdf, lb_file_b,
                                            var_status, lb_col, root)).pack(side="left", padx=(4, 0))
    # 对比模式下显示 rC，隐藏 rB；并动态修改 A/B 的标签
    roots = {"A": lab_a, "B": lab_b, "C": lab_c}
    for k, w in roots.items():
        w.txt = w.cget("text")

    def _refresh():
        t = lb_compare.get()
        _bind_a()
        _bind_b()
        if t == "BOM vs PDF":
            lab_a.configure(text="BOM Excel")
            lab_b.configure(text="原理图 PDF")
            rC.pack_forget()
            rB.pack(in_=parent, fill="x", pady=3)
        elif t == "PDF → Excel(器件核对)":
            lab_a.configure(text="原理图 PDF (基准)")
            lab_b.configure(text="BOM Excel (对照)")
            rC.pack_forget()
            rB.pack(in_=parent, fill="x", pady=3)
        else:
            lab_a.configure(text="文件 A")
            lab_b.configure(text="文件 B (PDF)" if t == "PDF vs PDF" else "文件 B (Excel)")
            if t == "PDF vs PDF":
                # 两个 PDF：行B隐藏，行C显示
                rB.pack_forget()
                lab_c.configure(text="文件 B (PDF)")
                rC.pack(in_=parent, fill="x", pady=3)
        # A 在对比模式下的文本
        if t == "Excel vs Excel":
            lab_a.configure(text="文件 A (Excel)")
        elif t == "PDF vs PDF":
            lab_a.configure(text="文件 A (PDF)")

    lb_compare.trace("w", lambda *a: _refresh())
    parent._refresh_files = _refresh
    _refresh()


def _browse_file(app, slot, lb_compare, lb_xlsx, lb_pdf, lb_file_b, var_status, lb_col, root):
    t = lb_compare.get()
    d = os.path.dirname(os.path.abspath(__file__))
    if t == "Excel vs Excel":
        p = filedialog.askopenfilename(title="选择 Excel", initialdir=d,
                                       filetypes=[("Excel", "*.xlsx"), ("所有文件", "*.*")])
        if not p:
            return
        if slot == "A":
            app["bom_path"] = p
            lb_xlsx.set(os.path.basename(p))
            try:
                headers, data = load_bom(p)
                app["headers"], app["data"] = headers, data
                if "Part Reference" in headers:
                    lb_col.set("Part Reference")
                _refresh_preview(app, lb_col)
            except Exception as e:
                messagebox.showerror("错误", "Excel A 读取失败:\n%s" % e, parent=root)
                return
        else:
            app["file_b"] = p
            lb_file_b.set(os.path.basename(p))
            try:
                headers_b, data_b = load_bom(p)
                app["headers_b"], app["data_b"] = headers_b, data_b
            except Exception as e:
                messagebox.showerror("错误", "Excel B 读取失败:\n%s" % e, parent=root)
                return
        var_status.set("已选 Excel：A=%s，B=%s" % (os.path.basename(app["bom_path"]),
                                                os.path.basename(app["file_b"])))
        return

    if t == "PDF vs PDF":
        p = filedialog.askopenfilename(title="选择 PDF", initialdir=d,
                                       filetypes=[("PDF", "*.pdf"), ("所有文件", "*.*")])
        if not p:
            return
        if slot == "A":
            app["pdf_path"] = p
            lb_pdf.set(os.path.basename(p))
        else:
            app["file_b"] = p
            lb_file_b.set(os.path.basename(p))
        var_status.set("已选 PDF：A=%s，B=%s" % (os.path.basename(app["pdf_path"]),
                                                os.path.basename(app["file_b"])))
        return

    if t == "PDF → Excel(器件核对)":
        # A = 原理图 PDF（基准），B = BOM Excel（对照）
        if slot == "A":
            p = filedialog.askopenfilename(title="选择原理图 PDF", initialdir=d,
                                           filetypes=[("PDF", "*.pdf"), ("所有文件", "*.*")])
            if p:
                app["pdf_path"] = p
                lb_pdf.set(os.path.basename(p))
                var_status.set("已选原理图 PDF：%s" % os.path.basename(p))
        else:
            p = filedialog.askopenfilename(title="选择 BOM Excel", initialdir=d,
                                           filetypes=[("Excel", "*.xlsx"), ("所有文件", "*.*")])
            if p:
                app["bom_path"] = p
                lb_xlsx.set(os.path.basename(p))
                try:
                    headers, data = load_bom(p)
                    app["headers"], app["data"] = headers, data
                    if lb_col is not None and "Part Reference" in headers:
                        lb_col.set("Part Reference")
                    _refresh_preview(app, lb_col)
                    var_status.set("已载入 BOM Excel：%s，共 %d 行" % (os.path.basename(p), len(data)))
                except Exception as e:
                    messagebox.showerror("错误", "BOM Excel 读取失败:\n%s" % e, parent=root)
        return

    # BOM vs PDF
    if slot == "A":
        _browse("BOM", app, root, var_status, lb_col)
    else:
        _browse("PDF", app, root, var_status, lb_col)


def _browse(kind, app, root, var_status, lb_col):
    d = os.path.dirname(os.path.abspath(__file__))
    if kind == "BOM":
        p = filedialog.askopenfilename(title="选择 BOM Excel", initialdir=d,
                                       filetypes=[("Excel", "*.xlsx"), ("所有文件", "*.*")])
        if p:
            app["bom_path"] = p
            try:
                headers, data = load_bom(p)
                app["headers"], app["data"] = headers, data
                if lb_col is not None and "Part Reference" in headers:
                    lb_col.set("Part Reference")
                _refresh_preview(app, lb_col)
                var_status.set("已载入 BOM，共 %d 行数据" % len(data))
            except Exception as e:
                messagebox.showerror("错误", "BOM 读取失败:\n%s" % e, parent=root)
        return
    p = filedialog.askopenfilename(title="选择原理图 PDF", initialdir=d,
                                   filetypes=[("PDF", "*.pdf"), ("所有文件", "*.*")])
    if p:
        app["pdf_path"] = p
        var_status.set("已选择 PDF：%s" % os.path.basename(p))


def _refresh_preview(app, lb_col):
    tree = app.get("preview_tree")
    if not tree or not app["headers"]:
        return
    tree["columns"] = app["headers"]
    for h in app["headers"]:
        tree.heading(h, text=h)
        tree.column(h, width=100, anchor="w")
    for i in tree.get_children():
        tree.delete(i)
    for row in app["data"][:10]:
        tree.insert("", "end", values=[str(v)[:60] for v in row["values"]])
    if lb_col is not None and "Part Reference" in app["headers"]:
        lb_col.set("Part Reference")
    # 通知"核对设置"刷新列下拉（核对列 / Excel列）
    for cb in (app.get("_col_refresh_callbacks") or []):
        try:
            cb()
        except Exception:
            pass


# ---------- 核对设置 ----------
def _setting_rows(parent, lb_col, lb_mode, app, lb_xlsx, lb_pdf, lb_file_b, lb_compare,
                  var_status, root):
    # 核对列：可编辑，默认使用当前 Excel 的 Bit 号列
    r1 = ttk.Frame(parent, style="Card.TFrame")
    r1.pack(fill="x", pady=3)
    ttk.Label(r1, text="核对列", style="Sub.TLabel", width=11, anchor="w").pack(side="left")
    cmb_col = ttk.Combobox(r1, textvariable=lb_col, width=28)
    cmb_col.pack(side="left", fill="x", expand=True)

    # Excel列：显示文件A 实际有哪些列，选择作为主键/搜索值来源的列
    r1b = ttk.Frame(parent, style="Card.TFrame")
    r1b.pack(fill="x", pady=3)
    ttk.Label(r1b, text="Excel列", style="Sub.TLabel", width=11, anchor="w").pack(side="left")
    lb_col_b = tk.StringVar()
    cmb_col2 = ttk.Combobox(r1b, textvariable=lb_col_b, width=28)
    cmb_col2.pack(side="left", fill="x", expand=True)

    def _fill_columns(*_):
        """把文件A的列名填进两个下拉，并自动选中 'Part Reference'（若无则第0列）。"""
        headers = app.get("headers") or []
        opts = list(headers)
        cmb_col["values"] = opts
        cmb_col2["values"] = opts
        if opts:
            cur = lb_col.get()
            if cur not in opts:
                lb_col.set("Part Reference" if "Part Reference" in opts else opts[0])
            cur2 = lb_col_b.get()
            if cur2 not in opts:
                lb_col_b.set(lb_col.get())

    cmb_col2.bind("<<ComboboxSelected>>", lambda e: lb_col.set(lb_col_b.get()))

    # 文件变化/类型变化时刷新列名
    _fill_columns()
    cbs = app.get("_col_refresh_callbacks") or []
    if not isinstance(cbs, list):
        cbs = []
    cbs.append(_fill_columns)
    app["_col_refresh_callbacks"] = cbs
    lb_xlsx.trace("w", _fill_columns)
    lb_compare.trace("w", _fill_columns)

    r2 = ttk.Frame(parent, style="Card.TFrame")
    r2.pack(fill="x", pady=3)
    ttk.Label(r2, text="匹配方式", style="Sub.TLabel", width=11, anchor="w").pack(side="left")
    cmb_mode = ttk.Combobox(r2, textvariable=lb_mode, width=28, state="readonly")
    cmb_mode.pack(side="left", fill="x", expand=True)
    cmb_mode["values"] = ["位号核对(推荐)", "全文匹配", "位号对比"]
    cmb_mode.current(0)

    # 主键比对方式：仅在 Excel vs Excel 模式下生效
    r4 = ttk.Frame(parent, style="Card.TFrame")
    r4.pack(fill="x", pady=3)
    ttk.Label(r4, text="主键比对", style="Sub.TLabel", width=11, anchor="w").pack(side="left")
    cmb_key = ttk.Combobox(r4, textvariable=app.setdefault("key_var", __import__("tkinter").StringVar()),
                           width=28, state="readonly")
    cmb_key.pack(side="left", fill="x", expand=True)
    cmb_key["values"] = ["A 逐位号比对(精细)", "B 按物料分组(聚合)"]
    cmb_key.current(0)

    # 比对字段选择（方式A用）
    r5 = ttk.Frame(parent, style="Card.TFrame")
    r5.pack(fill="x", pady=3)
    ttk.Label(r5, text="比对字段", style="Sub.TLabel", width=11, anchor="w").pack(side="left")
    cmb_fields = ttk.Combobox(r5, textvariable=app.setdefault("fields_var", __import__("tkinter").StringVar()),
                              width=28)
    cmb_fields.pack(side="left", fill="x", expand=True)
    cmb_fields["values"] = ["Value,Quantity,Manufacturer PN",
                            "Value,Quantity",
                            "Value,PCB Footprint,Quantity,Manufacturer PN",
                            "Value,PCB Footprint,Manufacturer PN"]
    cmb_fields.current(0)

    # 分组键选择（方式B用）
    r6 = ttk.Frame(parent, style="Card.TFrame")
    r6.pack(fill="x", pady=3)
    ttk.Label(r6, text="分组键", style="Sub.TLabel", width=11, anchor="w").pack(side="left")
    cmb_grp = ttk.Combobox(r6, textvariable=app.setdefault("grp_var", __import__("tkinter").StringVar()),
                           width=28)
    cmb_grp.pack(side="left", fill="x", expand=True)
    cmb_grp["values"] = ["Value,Manufacturer PN",
                         "Value,PCB Footprint",
                         "Value,Manufacturer PN,PCB Footprint",
                         "Value"]
    cmb_grp.current(0)

    # 根据对比类型显隐主键比对设置
    def _sync_key_rows(*_):
        ctype = lb_compare.get()
        if ctype == "Excel vs Excel":
            r4.pack(in_=parent, fill="x", pady=3, before=r2 if r2.winfo_ismapped() else None)
            r5.pack(in_=parent, fill="x", pady=3, before=r3)
            r6.pack(in_=parent, fill="x", pady=3, before=r3)
        else:
            r4.pack_forget()
            r5.pack_forget()
            r6.pack_forget()
    # 保持顺序：r1,r1b,r2,r4,r5,r6,r3
    r4.pack_forget(); r5.pack_forget(); r6.pack_forget()
    lb_compare.trace("w", _sync_key_rows)

    def _sync_key_var(*_):
        v = cmb_key.get()
        app["key_style"] = "B" if v.startswith("B") else "A"
    cmb_key.bind("<<ComboboxSelected>>", _sync_key_var)
    cmb_key.bind("<<ComboboxSelected>>", lambda e: _sync_key_var(e), add="+")

    def _sync_fields(*_):
        s = cmb_fields.get()
        if s:
            app["detail_fields"] = [x.strip() for x in s.split(",") if x.strip()]
    cmb_fields.bind("<FocusOut>", _sync_fields)

    def _sync_grp(*_):
        s = cmb_grp.get()
        if s:
            app["group_fields"] = [x.strip() for x in s.split(",") if x.strip()]
    cmb_grp.bind("<FocusOut>", _sync_grp)

    # 默认值
    app["key_style"] = "A"
    app["detail_fields"] = ["Value", "Quantity", "Manufacturer PN"]
    app["group_fields"] = ["Value", "Manufacturer PN"]

    r3 = ttk.Frame(parent, style="Card.TFrame")
    r3.pack(fill="x", pady=(8, 2))
    ttk.Button(r3, text=" 开始核对 ", style="Primary.TButton", command=lambda: threading.Thread(
        target=_do_check, args=(app, lb_xlsx, lb_pdf, lb_file_b, lb_compare, lb_col, lb_mode,
                                var_status, root),
        daemon=True).start()).pack(side="left", fill="x", expand=True)
    ttk.Button(r3, text="导出报告", style="Success.TButton", command=lambda: threading.Thread(
        target=_do_export, args=(app, root, var_status), daemon=True).start())\
        .pack(side="left", padx=(6, 0), fill="x", expand=True)


# ---------- 二级窗口按钮 ----------
def _subwin_rows(parent, root, app):
    btns = [
        ("PDF 文字查看", lambda: open_pdf_text_win(root, app)),
        ("不一致清单", lambda: open_mismatch_win(root, app)),
        ("汇总统计", lambda: open_summary_win(root, app)),
    ]
    for text, cmd in btns:
        ttk.Button(parent, text=text, style="TButton", command=cmd).pack(fill="x", pady=2)


# ---------- 预览卡片 ----------
def _preview_card(parent, app, lb_col):
    card = _make_card(parent, "BOM 数据预览")
    card.pack(fill="both", expand=True, pady=(0, 8))
    frame = ttk.Frame(card, style="Card.TFrame")
    frame.pack(fill="both", expand=True)
    tree = ttk.Treeview(frame, show="headings", height=6)
    vs = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
    hs = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=vs.set, xscrollcommand=hs.set)
    tree.pack(side="left", fill="both", expand=True)
    vs.pack(side="right", fill="y")
    hs.pack(side="bottom", fill="x")
    app["preview_tree"] = tree


def _result_card(parent, root, app):
    card = _make_card(parent, "核对结果  —  双击行查看位号详情")
    card.pack(fill="both", expand=True)
    frame = ttk.Frame(card, style="Card.TFrame")
    frame.pack(fill="both", expand=True)
    tree = ttk.Treeview(frame, columns=("row", "col", "des", "found", "pages", "raw"),
                        show="headings", height=18)
    for c, (t, w) in {
        "row": ("BOM行", 60), "col": ("列", 60), "des": ("位号/值", 110),
        "found": ("结果", 74), "pages": ("PDF页码", 90), "raw": ("原值(BOM)", 260),
    }.items():
        tree.heading(c, text=t)
        tree.column(c, width=w, anchor="w")
    vs = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
    hs = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=vs.set, xscrollcommand=hs.set)
    tree.pack(side="left", fill="both", expand=True)
    vs.pack(side="right", fill="y")
    hs.pack(side="bottom", fill="x")
    tree.tag_configure("ok", background=Theme.OK_BG, foreground=Theme.OK_FG)
    tree.tag_configure("bad", background=Theme.BAD_BG, foreground=Theme.BAD_FG)
    tree.tag_configure("odd", background="#FAFCFF")

    def on_double(event):
        # 对比模式下列结构不同，无需详情窗口
        cols = tree["columns"]
        if "des" not in cols:
            return
        sel = tree.selection()
        if not sel:
            return
        vals = tree.item(sel[0], "values")
        if not vals or len(vals) < 6:
            return
        row, col, des, found, pages, raw = vals
        pg = [int(x) for x in str(pages).split(",") if x.strip().isdigit()] if pages else []
        open_detail_win(root, app, des, pg, row, col, raw)

    tree.bind("<Double-1>", on_double)
    app["result_tree"] = tree


# ---------- 核对线程 ----------
def _do_check(app, lb_xlsx, lb_pdf, lb_file_b, lb_compare, lb_col, lb_mode, var_status, root):
    try:
        ctype = lb_compare.get()
        # 补齐路径
        _sync_paths(app, lb_xlsx, lb_pdf, lb_file_b, lb_compare)

        if ctype == "PDF → Excel(器件核对)":
            if not app["pdf_path"] or not app["bom_path"]:
                root.after(0, lambda: messagebox.showwarning("提示", "请选择 原理图PDF(A) 与 BOM Excel(B)", parent=root))
                return
            root.after(0, lambda: var_status.set("正在以 PDF 位号为准核对 Excel..."))
            rows, stats = compare_pdf_to_excel(app["pdf_path"], app["bom_path"],
                                               lb_col.get() or "Part Reference")
            app["result_rows"] = rows
            app["stats"] = stats
            oa = ob = 0
            root.after(0, lambda: _fill_compare_results(app, root, var_status, oa, ob))
            return

        if ctype == "Excel vs Excel":
            if not app["bom_path"] or not app["file_b"]:
                root.after(0, lambda: messagebox.showwarning("提示", "请选择 文件A 与 文件B (两个Excel)", parent=root))
                return
            # 主键比对方式：A 逐位号精细 / B 按物料分组
            key_style = app.get("key_style", "A")
            root.after(0, lambda: var_status.set(
                "正在按 %s 对比两个 Excel..." % ("物料分组" if key_style == "B" else "位号(主键)精细化")))
            if key_style == "B":
                gf = app.get("group_fields") or ["Value", "Manufacturer PN"]
                rows, stats = compare_excel_group(app["bom_path"], app["file_b"], tuple(gf))
            else:
                fields = app.get("detail_fields") or ["Value", "Quantity", "Manufacturer PN"]
                rows, stats = compare_excel_detail(app["bom_path"], app["file_b"],
                                                   lb_col.get() or "Part Reference", tuple(fields))
            app["result_rows"] = rows
            app["stats"] = stats
            oa, ob = len(stats["only_a"]), len(stats["only_b"])
            root.after(0, lambda: _fill_compare_results(app, root, var_status, oa, ob))
            return

        if ctype == "PDF vs PDF":
            if not app["pdf_path"] or not app["file_b"]:
                root.after(0, lambda: messagebox.showwarning("提示", "请选择 文件A 与 文件B (两个PDF)", parent=root))
                return
            root.after(0, lambda: var_status.set("正在解码两个 PDF..."))
            rows, stats = compare_pdf_pdf(app["pdf_path"], app["file_b"])
            app["result_rows"] = rows
            app["stats"] = stats
            oa, ob = len(stats["only_a"]), len(stats["only_b"])
            root.after(0, lambda: _fill_compare_results(app, root, var_status, oa, ob))
            return

        # ---- BOM vs PDF（原有逻辑）----
        if not app["bom_path"]:
            if lb_xlsx.get():
                app["bom_path"] = lb_xlsx.get()
            else:
                root.after(0, lambda: messagebox.showwarning("提示", "请选择 BOM 文件"))
                return
        if not app["pdf_path"]:
            if lb_pdf.get():
                app["pdf_path"] = lb_pdf.get()
            else:
                root.after(0, lambda: messagebox.showwarning("提示", "请选择 PDF 文件"))
                return

        root.after(0, lambda: var_status.set("正在解码 PDF 原理图文字..."))

        # 预热：若未载入 BOM
        if not app["headers"]:
            try:
                headers, data = load_bom(app["bom_path"])
                app["headers"], app["data"] = headers, data
            except Exception as e:
                err = str(e)
                root.after(0, lambda: messagebox.showerror("错误", "BOM 读取失败:\n%s" % err))
                return

        pages_words, pages_text, _ = decode_pdf_blocks(app["pdf_path"])
        app["pdf_pages_text"] = pages_text
        pdf_tokens, all_words = extract_designators(pages_words)
        app["pdf_tokens"] = pdf_tokens
        app["pdf_all_words"] = all_words

        # 位号上下文（供详情窗口）
        ctx = {}
        for pno, pwords in pages_words.items():
            for y, words in pwords:
                for w, x0, x1, top in words:
                    up = w.upper()
                    if re.match(r"^[A-Za-z]{1,4}\d{1,4}$", up):
                        ctx.setdefault(up, []).append(f"第{pno}页  y≈{y:.0f}   {w}")
        app["designator_context"] = ctx

        col_idx = app["headers"].index(lb_col.get()) if lb_col.get() in app["headers"] else 0
        mode = lb_mode.get()
        if "位号" in mode and "对比" not in mode:
            rows = check_designators(app["headers"], app["data"], col_idx, pdf_tokens)
        else:
            rows = check_value_in_pdf(app["headers"], app["data"], col_idx, all_words,
                                      designators_only=False)

        total, okc, badc, bad_list = CompareResult.stat(rows)
        all_bom = set()
        for r in rows:
            for d in r["designators"]:
                all_bom.add(d["designator"])
        extra = sorted(set(pdf_tokens.keys()) - all_bom)
        stats = {
            "all": {
                "BOM位号总数": total,
                "一致(在PDF中找到)": okc,
                "不一致(未找到)": badc,
                "PDF识别位号总数": len(pdf_tokens),
                "PDF中有而BOM无": len(extra),
            },
            "bad": [dict(b, row=r["header_row"]) for r in rows for b in r["designators"]
                    if not b["found"]],
            "extra": extra,
            "in_bom": all_bom,
        }
        app["result_rows"] = rows
        app["stats"] = stats

        root.after(0, lambda: _fill_results(app, root, var_status, total, okc, badc, len(extra)))
    except Exception as e:
        import traceback
        traceback.print_exc()
        err = str(e)
        root.after(0, lambda: messagebox.showerror("错误", "核对失败:\n%s" % err))


def _sync_paths(app, lb_xlsx, lb_pdf, lb_file_b, lb_compare):
    """把界面输入同步进 app。"""
    if lb_xlsx.get() and not app["bom_path"]:
        app["bom_path"] = lb_xlsx.get()
    if lb_pdf.get() and not app["pdf_path"]:
        app["pdf_path"] = lb_pdf.get()
    if lb_file_b.get() and not app["file_b"]:
        app["file_b"] = lb_file_b.get()


def _fill_compare_results(app, root, var_status, oa, ob):
    tree = app.get("result_tree")
    if not tree:
        return
    for i in tree.get_children():
        tree.delete(i)
    mode = app.get("stats", {}).get("mode", "set")
    if mode == "pdf2excel":
        tree["columns"] = ("item", "status", "valueA", "footA", "near")
        for c, (t, w) in {"item": ("PDF位号", 100), "status": ("核对结果", 130),
                          "valueA": ("Excel值(Value)", 130), "footA": ("Excel封装(Footprint)", 160),
                          "near": ("PDF附近标注", 340)}.items():
            tree.heading(c, text=t)
            tree.column(c, width=w, anchor="w")
        for row in app["result_rows"]:
            st = row["status"]
            tag = "ok" if st == "一致" else "bad"
            tree.insert("", "end",
                        values=(row["item"], st, row["valueA"], row["footA"], row["near"]),
                        tags=(tag,))
        n_ok = app.get("stats", {}).get("all", {}).get("值疑似一致", 0)
        n_may = app.get("stats", {}).get("all", {}).get("值疑似不一致/待确认", 0)
        n_miss = app.get("stats", {}).get("all", {}).get("PDF有Excel无", 0)
        var_status.set(f"器件核对完成：值一致 {n_ok}，待确认 {n_may}，PDF有Excel无 {n_miss}")
        messagebox.showinfo("器件核对完成",
                            f"值疑似一致: {n_ok}\n值疑似不一致/待确认: {n_may}\nPDF有Excel无: {n_miss}",
                            parent=root)
        return
    if mode == "detail":
        # 逐位号：列= 位号 | 结果 | 不一致字段详情
        tree["columns"] = ("item", "status", "fields")
        for c, (t, w) in {"item": ("位号(主键)", 120), "status": ("结果", 110),
                          "fields": ("字段差异(字段A→B)", 460)}.items():
            tree.heading(c, text=t)
            tree.column(c, width=w, anchor="w")
        for row in app["result_rows"]:
            st = row["status"]
            tag = "ok" if st == "一致" else "bad"
            if row["field_diffs"]:
                desc = "；".join("%s: %s → %s" % (d["field"], d["a"] or "空", d["b"] or "空")
                                 for d in row["field_diffs"])
            else:
                desc = ""
            tree.insert("", "end", values=(row["item"], st, desc), tags=(tag,))
    elif mode == "group":
        # 按物料分组：列= 分组 | 结果 | 数量A | 数量B
        tree["columns"] = ("item", "status", "qa", "qb")
        for c, (t, w) in {"item": ("物料分组(Value | MPN)", 300), "status": ("结果", 110),
                          "qa": ("数量A", 60), "qb": ("数量B", 60)}.items():
            tree.heading(c, text=t)
            tree.column(c, width=w, anchor="w")
        for row in app["result_rows"]:
            st = row["status"]
            tag = "ok" if st == "一致" else "bad"
            tree.insert("", "end", values=(row["item"], st, row["qty_a"], row["qty_b"]),
                        tags=(tag,))
    else:
        tree["columns"] = ("item", "status", "a", "b")
        for c, (t, w) in {"item": ("器件/位号", 130), "status": ("结果", 90),
                          "a": ("文件A", 150), "b": ("文件B", 150)}.items():
            tree.heading(c, text=t)
            tree.column(c, width=w, anchor="w")
        for row in app["result_rows"]:
            st = row["status"]
            tag = "ok" if st == "一致" else "bad"
            tree.insert("", "end", values=(row["item"], st, row["a_pages"] or row["a"],
                                           row["b_pages"] or row["b"]), tags=(tag,))
    n_common = len(app.get("stats", {}).get("common", []))
    var_status.set(f"对比完成：一致 {n_common}，仅A有 {oa}，仅B有 {ob}")
    total = len(app["result_rows"])
    messagebox.showinfo("对比完成",
                        f"共 {total} 项\n 一致: {n_common}\n 仅A有: {oa}  仅B有: {ob}",
                        parent=root)


def _fill_results(app, root, var_status, total, okc, badc, nextra):
    tree = app.get("result_tree")
    if not tree:
        return
    for i in tree.get_children():
        tree.delete(i)
    odd = False
    for row in app["result_rows"]:
        for d in row["designators"]:
            ok = d["found"]
            odd = not odd
            tree.insert("", "end",
                        values=(row["header_row"], row["col_name"], d["designator"],
                                "一致" if ok else "不一致",
                                ",".join(str(p) for p in d["pages"]) if ok else "",
                                str(row["raw"])[:90]),
                        tags=("ok" if ok else "bad",)) if ok else tree.insert(
                "", "end",
                values=(row["header_row"], row["col_name"], d["designator"], "不一致",
                        "", str(row["raw"])[:90]),
                tags=("bad",))
    var_status.set(f"核对完成：共 {total} 个位号，一致 {okc}，不一致 {badc}，PDF额外位号 {nextra}")
    messagebox.showinfo("核对完成",
                        f"共 {total} 个位号\n一致: {okc}\n不一致: {badc}\nPDF额外位号: {nextra}",
                        parent=root)


# ---------- 导出报告 ----------
def _do_export(app, root, var_status):
    try:
        if not app.get("stats"):
            root.after(0, lambda: messagebox.showwarning("提示", "请先执行“开始核对”。", parent=root))
            return
        d = os.path.dirname(os.path.abspath(__file__))
        xlsx_path, txt_path = gen_report(app["bom_path"], app["result_rows"], app["stats"],
                                         app["pdf_tokens"], app["pdf_pages_text"], d)
        msg = "报告已生成:\n%s\n%s" % (xlsx_path, txt_path)
        root.after(0, lambda: (var_status.set(msg), messagebox.showinfo("完成", msg, parent=root)))
    except Exception as e:
        traceback.print_exc()
        err = str(e)
        root.after(0, lambda: messagebox.showerror("错误", "报告生成失败:\n%s" % err, parent=root))


# ---------- 自动载入 ----------
def _auto_load(app, lb_xlsx, lb_pdf, lb_col):
    import glob
    d = os.path.dirname(os.path.abspath(__file__))
    xlsx = glob.glob(os.path.join(d, "*.BOM.xlsx")) + glob.glob(os.path.join(d, "*BOM*.xlsx"))
    pdf = glob.glob(os.path.join(d, "*.pdf"))
    if xlsx:
        p = sorted(xlsx)[0]
        app["bom_path"] = p
        lb_xlsx.set(os.path.basename(p))
        try:
            headers, data = load_bom(p)
            app["headers"], app["data"] = headers, data
        except Exception:
            pass
    if pdf:
        p = sorted(pdf)[0]
        app["pdf_path"] = p
        lb_pdf.set(os.path.basename(p))
    _refresh_preview(app, lb_col)


# ============================================================
#  命令行入口（GUI / 命令行核对 / 帮助 三模式）
# ============================================================
def run_cli(bom_path=None, pdf_path=None, col_name="Part Reference", mode="位号核对(推荐)",
            out_dir=None, file_b=None, ctype=None):
    """命令行核对/对比模式：无需图形界面即可输出结果并生成报告。
    ctype: BOM vs PDF / Excel vs Excel / PDF vs PDF
    """
    # 对比模式：两个同类文件
    if ctype in ("Excel vs Excel", "PDF vs PDF"):
        if not bom_path or not file_b:
            print("用法: --check --compare 类型 --a <文件A> --b <文件B> [--col 列名]")
            return 1
        if ctype == "Excel vs Excel":
            print("[1/3] 对比两个 Excel 位号...")
            rows, stats = compare_excel_excel(bom_path, file_b, col_name)
        else:
            print("[1/3] 解码并对比两个 PDF...")
            rows, stats = compare_pdf_pdf(bom_path, file_b)
        oa, ob = len(stats["only_a"]), len(stats["only_b"])
        print("[2/3] 结果: 两文件一致 %d  仅A有 %d  仅B有 %d" % (len(stats["common"]), oa, ob))
        print("[3/3] 生成报告...")
        if not out_dir:
            out_dir = os.path.dirname(os.path.abspath(bom_path))
        xlsx_path, txt_path = gen_report(bom_path, rows, stats, {}, {}, out_dir)
        print("报告已生成:")
        print("   %s" % xlsx_path)
        print("   %s" % txt_path)
        return 0 if oa == 0 and ob == 0 else 2

    # BOM vs PDF 核对（原有逻辑）
    if not bom_path or not pdf_path:
        print("用法: bom_pdf_verify.py --check --bom <BOM.xlsx> --pdf <原理图.pdf> [--col 列名]")
        print("      或双击运行进入 GUI 界面")
        return 1
    headers, data = load_bom(bom_path)
    if col_name not in headers:
        print("错误: 列名 '%s' 不存在。可用列: %s" % (col_name, ", ".join(headers)))
        return 1
    print("[1/4] 解码 PDF 原理图文字...")
    pages_words, pages_text, _ = decode_pdf_blocks(pdf_path)
    pdf_tokens, all_words = extract_designators(pages_words)
    col_idx = headers.index(col_name)
    if "位号" in mode:
        rows = check_designators(headers, data, col_idx, pdf_tokens)
    else:
        rows = check_value_in_pdf(headers, data, col_idx, all_words, designators_only=False)
    total, okc, badc, bad_list = CompareResult.stat(rows)
    all_bom = set()
    for r in rows:
        for d in r["designators"]:
            all_bom.add(d["designator"])
    extra = sorted(set(pdf_tokens.keys()) - all_bom)
    stats = {
        "all": {
            "BOM位号总数": total,
            "一致(在PDF中找到)": okc,
            "不一致(未找到)": badc,
            "PDF识别位号总数": len(pdf_tokens),
            "PDF中有而BOM无": len(extra),
        },
        "bad": [dict(b, row=r["header_row"]) for r in rows for b in r["designators"]
                if not b["found"]],
        "extra": extra,
        "in_bom": all_bom,
    }
    print("[2/4] 核对结果: 共 %d 个位号，一致 %d，不一致 %d" % (total, okc, badc))
    print("[3/4] 不一致(未在PDF中找到): %d" % len(bad_list))
    for b in bad_list:
        print("      MISS  %s  (BOM第%d行)" % (b["designator"], 0))
    print("[4/4] 生成报告...")
    if not out_dir:
        out_dir = os.path.dirname(os.path.abspath(bom_path))
    xlsx_path, txt_path = gen_report(bom_path, rows, stats, pdf_tokens, pages_text, out_dir)
    print("报告已生成:")
    print("   %s" % xlsx_path)
    print("   %s" % txt_path)
    return 0 if badc == 0 else 2


def main():
    import argparse
    parser = argparse.ArgumentParser(
        description="BOM(Excel) 与 PDF 原理图 器件核对 / 对比工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="示例:\n"
               "  双击 / 无参数      : 打开 GUI 界面（支持三种对比类型）\n"
               "  命令行核对(BOMvPDF): --check --bom xxx.xlsx --pdf xxx.pdf\n"
               "  双 Excel 对比      : --check --compare 'Excel vs Excel' --a a.xlsx --b b.xlsx\n"
               "  双 PDF 对比        : --check --compare 'PDF vs PDF' --a a.pdf --b b.pdf\n"
               "  指定列/模式        : 加 --col 'Value' --match 全文匹配 --out ./report\n")
    parser.add_argument("--gui", action="store_true", help="启动图形界面(默认)")
    parser.add_argument("--check", action="store_true", help="命令行模式(无需GUI)")
    parser.add_argument("--compare", type=str, default=None,
                        choices=["Excel vs Excel", "PDF vs PDF"], help="双文件对比模式")
    parser.add_argument("--a", type=str, default=None, help="文件A 路径")
    parser.add_argument("--b", type=str, default=None, help="文件B 路径(对比用)")
    parser.add_argument("--bom", type=str, default=None, help="BOM Excel (.xlsx) 路径")
    parser.add_argument("--pdf", type=str, default=None, help="原理图 PDF 路径")
    parser.add_argument("--col", type=str, default="Part Reference", help="要核对的列名")
    parser.add_argument("--match", type=str, default="位号核对(推荐)",
                        choices=["位号核对(推荐)", "全文匹配"], help="匹配方式")
    parser.add_argument("--out", type=str, default=None, help="报告输出目录(默认=A所在目录)")
    args = parser.parse_args()

    if args.compare:
        code = run_cli(args.a, args.pdf, args.col, args.match, args.out, args.b, args.compare)
        sys.exit(code)
    if args.check:
        code = run_cli(args.bom, args.pdf, args.col, args.match, args.out)
        sys.exit(code)
    # 默认进入 GUI
    run_gui()


if __name__ == "__main__":
    main()

